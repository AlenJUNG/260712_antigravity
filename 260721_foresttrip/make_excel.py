import json
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import re

def normalize_key(name):
    name = re.sub(r'^\([^)]*\)', '', name) # Remove prefix like (가평군)
    name = name.replace(' ', '')
    suffixes = [
        '자연휴양림', '휴양림', '숲속야영장', '캠핑랜드', '휴양랜드', 
        '산림휴양마을', '산림문화타운', '숲속문화촌', '산림레포츠파크', 
        '항노화힐링랜드', '쉬자파크', '해송', '폭포'
    ]
    for suffix in suffixes:
        if name.endswith(suffix):
            if len(name) > len(suffix):
                name = name[:-len(suffix)]
    return name

ADDITIONAL_OPEN_YEARS = {
    # Gyeonggi/Incheon
    '강씨봉': 2011, '칼봉산': 2008, '석모도': 2011, '축령산': 1995,
    '수락산동막골': 2025, '동두천': 2020, '서운산': 2018, '신암저수지': 2024,
    '양평백운봉': 2017, '양평쉬자': 2018, '쉬자파크': 2018, '고대산': 2017,
    '덕적도': 2025, '용인': 2009, '바라산': 2014, '천보산': 2012,
    '무봉산': 2023, '청평': 2001, '강화': 2011, '양평설매재': 1999, '설매재': 1999,
    '양평': 2018, '의왕바라산': 2014,
    # Gangwon
    '진부령': 2026, '임해': 2009, '광치': 2006, '송이밸리': 2012,
    '망경대산': 2012, '치악산': 1994, '갯골': 2023, '하추': 2008,
    '두루웰': 2017, '강원숲체험장': 1997, '집다리골': 1994, '춘천숲': 2008,
    '태백고원': 2005, '평창': 2012, '가리산': 1998, '삼척활기': 2020,
    '피노키오': 2005, '횡성': 2002, '철원두루웰': 2017,
    # Chungbuk
    '성불산': 2016, '조령산': 1995, '소백산': 2017, '소선암': 2005,
    '속리산숲체험': 2017, '충북알프스': 2010, '민주지산': 2004, '장령산': 1994,
    '백야': 2011, '수레의산': 2007, '박달재': 1992, '옥전': 2022,
    '좌구산': 2009, '생거진천': 2014, '미원별빛': 2026, '옥화': 1999,
    '계명산': 1997, '문성': 2008, '봉황': 1996, '속리산숲체험휴양마을': 2017,
    # Daejeon/Chungnam
    '공주산림휴양': 2016, '금산산림문화': 2008, '양촌': 2013, '만인산': 1990,
    '장태산': 1991, '성주산': 1993, '원산도': 2026, '만수산': 1992,
    '영인산': 1997, '봉수산': 2007, '태학산': 2001, '칠갑산': 1993,
    '안면도': 1992, '용봉산': 1993, '희리산': 1999, '공주': 2016, '금산': 2008,
    # Jeonbuk
    '선암': 2023, '김제선암': 2023, '무주향로산': 2018, '고산': 1998,
    '성수산왕의숲': 2024, '방화동': 2003, '와룡': 1996, '내장산': 2026,
    '데미샘': 2012, '남원': 1995,
    # Jeonnam/Gwangju
    '주작산': 2007, '팔영산': 1998, '산수유': 2015, '제암산': 1996,
    '순천': 2011, '봉황산': 2012, '기찬': 2023, '완도': 2018,
    '흑석산': 1999, '백아산': 1996, '한천': 2003, '무등산편백': 1997,
    # Daegu/Gyeongbuk
    '토함산': 1997, '미숭산': 2012, '옥성': 2007, '군위장곡': 1997,
    '수도산': 2014, '비슬산': 1998, '화원': 2010, '문수산': 2020,
    '성주봉': 2001, '독용산성': 2014, '안동호반': 2010, '영양에코둥지': 2014,
    '보현산': 2022, '운주산승마': 2009, '구수곡': 2001, '금봉': 2004,
    '청도': 2022, '청송': 1997, '송정': 2006, '팔공산금화': 2016,
    '비학산': 2015, '학가산우래': 2000, '신불산': 1989,
    # Busan/Gyeongnam
    '거제': 1993, '거창산림레포츠': 2025, '금원산': 1993, '항노화힐링': 2021,
    '갈모봉': 2024, '도래재': 2022, '사천케이블카': 2021, '산청한방': 2014,
    '대운산': 2009, '자굴산': 2022, '진주월아산': 2022, '화왕산': 2014,
    '구재봉': 2016, '하동편백': 2020, '대봉산': 2021, '대봉캠핑': 2021,
    '산삼': 2012, '용추': 1993, '오도산': 2002, '중산': 1997,
    '덕원': 2019, '거창': 2025, '항노화힐링랜드': 2021, '대봉': 2021,
    # Jeju
    '붉은오름': 2012, '서귀포': 1995, '교래': 2011, '제주절물': 1997
}

def make_excel():
    # Load raw data
    raw_path = 'data/all-forests-raw.json'
    if not os.path.exists(raw_path):
        print(f"Error: {raw_path} not found.")
        return

    with open(raw_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Load scraped forest features
    features_path = 'data/forest-features.json'
    features = {}
    if os.path.exists(features_path):
        with open(features_path, 'r', encoding='utf-8') as f:
            features = json.load(f)

    rows = []
    for item in data:
        sido_code = int(item['sidoCode'])
        sido_name = item['sidoName']
        instt_id = item['insttId']
        instt_nm = item['insttNm']
        instt_tp_cd = item['insttTpCd']
        instt_tp_nm = item['insttTpNm']
        
        max_nights = item.get('maxNights')
        cycle_type = item.get('cycleType')
        
        # Open year details
        opened = item.get('opened')
        designated = item.get('designated')
        built = item.get('built')
        location = item.get('location')
        
        # Fallback to internet searched opening years
        if opened is None:
            norm_key = normalize_key(instt_nm)
            opened = ADDITIONAL_OPEN_YEARS.get(norm_key)
            
        # New status: 6 years or less since opening is "신축", otherwise "구축"
        is_new = "-"
        if opened is not None:
            this_year = 2026 # Use 2026 as current year context
            is_new = "신축" if (this_year - opened <= 6) else "구축"
            
        opened_val = int(opened) if opened is not None else "-"
        designated_val = int(designated) if designated is not None else "-"
        built_val = int(built) if built is not None else "-"
        location_val = location if location else "-"
        max_nights_val = int(max_nights) if max_nights is not None else "-"
        cycle_type_val = cycle_type if cycle_type else "-"
        
        # Analyze specialDates
        special_dates = item.get('specialDates', [])
        holidays = []
        lottery_found = False
        
        weekdays_ko = ["월", "화", "수", "목", "금", "토", "일"]
        holiday_weekdays = set()
        
        for sd in special_dates:
            dt_str = sd.get('date') # YYYYMMDD
            code = sd.get('code')
            name = sd.get('name')
            
            # code 01 = 정기휴무일
            if code == '01' or '휴무' in name or '휴관' in name:
                try:
                    dt_obj = datetime.strptime(dt_str, '%Y%m%d')
                    holiday_weekdays.add(dt_obj.weekday())
                except Exception:
                    pass
            # code 02 = 주말추첨
            if code == '02' or '추첨' in name:
                lottery_found = True
                
        notes = []
        if holiday_weekdays:
            wd_names = [weekdays_ko[w] for w in sorted(list(holiday_weekdays))]
            notes.append(f"정기휴무({','.join(wd_names)}요일)")
        if lottery_found:
            notes.append("주말추첨제 운영")
            
        notes_val = ", ".join(notes) if notes else "특이사항 없음"
        
        # Lookup features
        feat = features.get(instt_nm, {"valley": "X", "water": "X"})
        water_val = feat.get("water", "X")
        valley_val = feat.get("valley", "X")

        rows.append({
            'sido_code': sido_code,
            'instt_tp_cd': instt_tp_cd,
            '시도': sido_name,
            '구분': instt_tp_nm,
            '휴양림명': instt_nm,
            '기관 ID': instt_id,
            '개관연도': opened_val,
            '신축여부': is_new,
            '지정연도': designated_val,
            '조성연도': built_val,
            '최대숙박일수': max_nights_val,
            '예약주기': cycle_type_val,
            '주소/위치': location_val,
            '운영특이사항': notes_val,
            '물놀이': water_val,
            '계곡': valley_val
        })

    # Create DataFrame
    df = pd.DataFrame(rows)
    # Sort by Sido code first, then institution type code, then forest name
    df = df.sort_values(by=['sido_code', 'instt_tp_cd', '휴양림명'])
    df = df.drop(columns=['sido_code', 'instt_tp_cd'])

    # Output path
    output_xlsx = '전국휴양림정보.xlsx'

    # Save to path with styling
    for path in [output_xlsx]:
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='전국휴양림정보', index=False)
            
            # Format and Style sheet
            workbook = writer.book
            worksheet = writer.sheets['전국휴양림정보']
            
            # Ensure grid lines are visible
            worksheet.views.sheetView[0].showGridLines = True
            
            # Fonts and styles
            font_family = '맑은 고딕'
            header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Slate Blue
            
            body_font = Font(name=font_family, size=10)
            zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            thin_border_side = Side(border_style='thin', color='D9D9D9')
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')
            
            # Formats for columns
            # Column mapping index (1-based for Excel)
            # 1: 시도, 2: 구분, 3: 휴양림명, 4: 기관 ID, 5: 개관연도, 6: 신축여부, 7: 지정연도, 8: 조성연도, 9: 최대숙박일수, 10: 예약주기, 11: 주소/위치, 12: 운영특이사항, 13: 물놀이, 14: 계곡
            alignments = {
                1: align_center,  # 시도
                2: align_center,  # 구분
                3: align_left,    # 휴양림명
                4: align_center,  # 기관 ID
                5: align_center,  # 개관연도
                6: align_center,  # 신축여부
                7: align_center,  # 지정연도
                8: align_center,  # 조성연도
                9: align_center,  # 최대숙박일수
                10: align_center, # 예약주기
                11: align_left,    # 주소/위치
                12: align_left,    # 운영특이사항
                13: align_center,  # 물놀이
                14: align_center   # 계곡
            }
            
            # Format header row
            worksheet.row_dimensions[1].height = 28
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
                
            # Format data rows
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 20
                is_zebra = (row_idx % 2 == 0)
                row_fill = zebra_fill if is_zebra else white_fill
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = body_font
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = alignments.get(col_idx, align_left)
                    
                    # Number format if integer
                    if isinstance(cell.value, int):
                        cell.number_format = '#,##0'
                        
            # Auto-fit columns with safety margin
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    # Calculate visual width (Korean chars take more space)
                    visual_len = 0
                    for char in val:
                        if ord(char) > 127: # non-ASCII (Korean, etc.)
                            visual_len += 2
                        else:
                            visual_len += 1
                    if visual_len > max_len:
                        max_len = visual_len
                # Add padding
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        print(f"Excel file successfully saved to: {path}")

if __name__ == '__main__':
    make_excel()
